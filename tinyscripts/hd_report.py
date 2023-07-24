import warnings
import openpyxl
import os
import sys
from datetime import datetime, timedelta, time, date
from string import digits
import csv

from tkinter import Tk, Frame, Button, Menu, filedialog, BooleanVar, Checkbutton, Radiobutton, ttk
from tkinter import messagebox as mbox
from lib import l


# ФИО можно посмотреть, набрав в адресной строке:  https://finfort.planfix.ru/?action=logincard&id=5316170
CONVERSION = {
    'ТП 1 линия': 'ТП₁',
    'ТП Финброкер': 'ТПфб',
    '_Запросы на техподдержку': 'TПit',
    'ТП 1 линия-Работа с Поставщиком': 'ТПvndr',
    'ТП₁-Работа с Поставщиком': 'ТПvndr',
    'Заявки для РКО': 'ТПрко',
    'Вопросы ПланФикса': 'ТПпф'
}

MAX_BLANK_LINES = 5
START_WORKTIME = time(8,0,0)
END_WORKTIME_8 = time(17,0,0)
END_WORKTIME_12 = time(20,0,0)
HOURS_8 = timedelta(hours=9)
HOURS_12 = timedelta(hours=12)

class Example(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.parent = parent
        self.initUI()

    def initUI(self):
        self.parent.title("Отчет всех линий ТП v 0.17 от 12.07.2023")
        self.pack()
        # Главное меню .
        menubar = Menu(self.parent, font=('Calibri 10'))
        self.parent.config(menu=menubar)
        # Раздел меню, без открепления
        self.fileMenu = Menu(menubar, tearoff=0, font=('Calibri 10'))
        # Пункты раздела меню
        self.fileMenu.add_command(label="Открыть отчет из ПланФикса (Excel)", command=self.onOpen)
        self.fileMenu.add_separator()
        self.fileMenu.add_command(label="Выход", command=self.onExit)
        # Присоединяем раздел каскадом
        menubar.add_cascade(label="Файл", menu=self.fileMenu)

        btnOpen = Button(self, text="Открыть отчет из ПланФикса (Excel)", command=self.onOpen, font=('Calibri 10'))
        btnOpen.grid(padx=5, pady=5, row=0, column=0)

        self.progressBar = ttk.Progressbar(self, orient="horizontal", mode="determinate", maximum=100, value=0,
                                           length=200)
        self.progressBar.grid(padx=5, pady=5, row=1, column=0)#, columnspan=2)
        self.xlsx = ''
        self.pf_users = {}
        try:
            with open('users.csv', 'r', encoding="utf-8") as csv_file:
                reader = csv.DictReader(csv_file)
                for row in reader:
                    self.pf_users[int(row['id'])] = row['fio']
                    q=0
        except Exception as e:
            mbox.showerror('Ошибка в файле users.csv', 'Запросите корректный файлв ТП ИТ и поместите его рядом с .exe')
            sys.exit()

    def onExit(self):
        #mbox.showerror('Ошибка в файле' +  self.xlsx, 'Нет столбца с... ')
        sys.exit()

    def workTime(self, start_datetime, end_datetime, alldays=False):
        """ Вычисляет период рабочего времени, прошедшего между start_datetime, end_datetime.
        работает в двух режимах: 5/2 с 8 до 17 (alldays=False) и 7/0 с 8 до 20 (alldays=True) """
        if end_datetime <= start_datetime:
            return timedelta(microseconds=1)
        s = start_datetime
        e = end_datetime
        s = datetime(s.timetuple().tm_year, s.timetuple().tm_mon, s.timetuple().tm_mday, 0, 0, 1)
        e = datetime(e.timetuple().tm_year, e.timetuple().tm_mon, e.timetuple().tm_mday, 23, 59, 59)
        count_of_days = round((e - s).total_seconds() / (3600 * 24))
        worktime = timedelta(microseconds=1)
        first_day_start_time = datetime.combine(date(start_datetime.timetuple().tm_year,
                                                     start_datetime.timetuple().tm_mon,
                                                     start_datetime.timetuple().tm_mday), START_WORKTIME)
        last_day_start_time = datetime.combine(date(end_datetime.timetuple().tm_year,
                                                    end_datetime.timetuple().tm_mon,
                                                    end_datetime.timetuple().tm_mday), START_WORKTIME)
        if alldays:                 # режим 7/0 с 8 до 20
            first_day_end_time = datetime.combine(date(start_datetime.timetuple().tm_year,
                                                       start_datetime.timetuple().tm_mon,
                                                       start_datetime.timetuple().tm_mday), END_WORKTIME_12)
            last_day_end_time = datetime.combine(date(end_datetime.timetuple().tm_year,
                                                      end_datetime.timetuple().tm_mon,
                                                      end_datetime.timetuple().tm_mday), END_WORKTIME_12)
        else:                       # режим 5/2 с 8 до 17
            first_day_end_time = datetime.combine(date(start_datetime.timetuple().tm_year,
                                                       start_datetime.timetuple().tm_mon,
                                                       start_datetime.timetuple().tm_mday), END_WORKTIME_8)
            last_day_end_time = datetime.combine(date(end_datetime.timetuple().tm_year,
                                                      end_datetime.timetuple().tm_mon,
                                                      end_datetime.timetuple().tm_mday), END_WORKTIME_8)

        for i in range(0, count_of_days):
            if s.timetuple().tm_wday > 4 and not alldays:     # Проверка на рабочий день
                break
            if start_datetime.date() == end_datetime.date():  # Если этап завершился за один день
                if end_datetime < first_day_start_time:  # Если закончилось раньше начала раб.дня
                    break
                if start_datetime > last_day_end_time:  # Если началось позже окончания раб.дня
                    break
                if start_datetime < first_day_start_time and end_datetime > last_day_end_time: # Если весь день
                    if alldays:
                        worktime += HOURS_12
                    else:
                        worktime += HOURS_8
                else:
                    if start_datetime < last_day_start_time:
                        start_datetime = last_day_start_time    # Если началось раньше начала раб.дня
                    if end_datetime > first_day_end_time:
                        end_datetime = first_day_end_time   # Если закончилось позже окончания раб.дня
                    worktime += end_datetime - start_datetime
            elif s.date() == start_datetime.date():         # Первый день многодневного этапа
                if start_datetime > first_day_end_time:     # Если началось позже окончания раб.дня
                    break
                if start_datetime < first_day_end_time:
                    if start_datetime < first_day_start_time:   # Если началось раньше начала раб.дня
                        if alldays:
                            worktime += HOURS_12
                        else:
                            worktime += HOURS_8
                    else:
                        worktime += first_day_end_time - start_datetime
            elif s.date() == end_datetime.date():           # Последний день многодневного этапа
                if end_datetime < last_day_start_time:      # Если закончилось раньше начала раб.дня
                    break
                if end_datetime > last_day_end_time:        # Если закончилось позже окончания раб.дня
                    if alldays:
                        worktime += HOURS_12
                    else:
                        worktime += HOURS_8
                else:
                    worktime += end_datetime - last_day_start_time
            else:                                           # Промежуточный день многодневного этапа
                if alldays:
                    worktime += HOURS_12
                else:
                    worktime += HOURS_8
            s += timedelta(days=1)
        return worktime

    def getPFuser(self,users):
        if l(users.split(',')[0]) and \
                str(users.split(',')[0]).find('Техподдержка(') == -1:
            # Когда несколько юзеров через запятую - берем первого
            if self.pf_users.get(l(users.split(',')[0])):
                return self.pf_users[int(users.split(',')[0])]
            else:
                return users.split(',')[0]
        else:
            return users  # нет - указываем всё поле

    def onOpen(self):
        self.progressBar['value'] = 0
        self.update()
        self.xlsx = ''
        dlg = filedialog.Open(self, filetypes=[('Файлы Excel 2007-2020', '*.xlsx')], initialdir='/home/da3/Загрузки')
        self.xlsx = dlg.show()
        tasks = {}
        if str(self.xlsx).find('.xlsx') > -1:   # Проверяем то, что и так понятно
            wb = openpyxl.load_workbook(filename=self.xlsx, read_only=True)
            ws = wb[wb.sheetnames[0]]
            wb_rez = openpyxl.Workbook(write_only=True)
            titles = []
            colnames = []
            blank_lines = 0
            self.progressBar['maximum'] = ws.max_row
            self.progressBar['value'] = 0
            self.update()
            ws_rez = wb_rez.create_sheet('Отчет работы всех линий ТП')
            ws_rez.append(['№ п/п', 'ID задачи', 'Название задачи', 'Постановщик', 'Стартовый проект', 'Создана',
                           'ТП₁ Новая(мин)', 'ТП₁ В работе(мин)', 'ТП₁ Исполнитель',
                           'ТПit Новая(мин)', 'ТПit В работе(мин)', 'ТПit Исполнитель',
                           'ТПvndr В работе(мин)', 'ТПvndr Исполнитель',
                           'ТПфб Новая(мин)', 'ТПфб В работе(мин)', 'ТПфб Исполнитель',
                           'ТПрко Новая(мин)', 'ТПрко В работе(мин)', 'ТПрко Исполнитель',
                           'Выполнено', 'Затраты ТП(мин)', 'Исполнитель (выполнено)', 'Завершено (факт)',
                           'Текущий статус', 'Эволюция статусов', 'Продукт', 'Канал продаж', 'Точка входа',
                           'Источник проблемы', 'Краткое описание проблемы', 'Создана - дата','Время'])
            for i, row in enumerate(ws.values):                 # Идем по строчкам отчета
                # print(i, row[0])
                if i == 0:
                    colnames = tuple(row)                       # Собираем имена колонок
                elif i > 0:
                    self.progressBar['value'] = i
                    self.update()
                    if not row[colnames.index('Лог')]:          # Если пустая колонка "Лог"
                        blank_lines += 1
                        if blank_lines > MAX_BLANK_LINES:
                            break
                    else:
                        stages_dic = {}
                        j = -1
                        last_project = ''
                        last_status = ''
                        last_datetime = datetime(1971,1,1,1,0,0)
                        last_user = ''
                        logs = row[colnames.index('Лог')]
                        last_delta = 0
                        evolution = ''
                        for log in logs.split('$'):             # $  разделяет строчки лога
                            if log and len(log) > 7:
                                # ~ разделяет поля лога fields: ДатаВремя[0] Проект[1] Статус[2] Исполнитель[3]
                                fileds = log.split('~')
                                if not last_status:             # в первый раз
                                    last_status = 'Новая'       # делаем дополнительный этап Новая перед всеми
                                    last_project = fileds[1]
                                    last_datetime = datetime.strptime(fileds[0].strip(), '%d-%m-%Y %H:%M')
                                    last_user = ''
                                    stages_dic[last_datetime] = {'project': last_project, 'status': last_status,
                                                                 'user': last_user}
                                if last_status == fileds[2] and last_project == fileds[1] and \
                                        datetime.strptime(fileds[0].strip(), '%d-%m-%Y %H:%M') >= last_datetime:
                                    last_datetime = datetime.strptime(fileds[0].strip(), '%d-%m-%Y %H:%M')
                                    pass                # Если дубль статуса и проекта и время >= , обновляем время
                                elif stages_dic.get(datetime.strptime(fileds[0].strip(), '%d-%m-%Y %H:%M'), None):
                                    # Если запись с таким временем уже была
                                    last_delta += 1 # Кол-во микросекунд, которое добавляем чтобы записи stages_dic
                                                    # с одинаковым временем не перезаписывали друг друга
                                    last_status = fileds[2]
                                    last_project = fileds[1]
                                    last_user = self.getPFuser(fileds[3])
                                    stages_dic[last_datetime + timedelta(microseconds=last_delta)] = {
                                        'project': last_project, 'status': last_status, 'user': last_user}
                                else:
                                    if last_datetime > datetime.strptime(fileds[0].strip(), '%d-%m-%Y %H:%M'):
                                        evolution = '⇵'       # Непоследовательное время(слияние двух задач или сбой)
                                    last_status = fileds[2]
                                    last_project = fileds[1]
                                    last_datetime = datetime.strptime(fileds[0].strip(), '%d-%m-%Y %H:%M')
                                    last_user = self.getPFuser(fileds[3])
                                    stages_dic[last_datetime] = {'project': last_project, 'status': last_status,
                                                                 'user': last_user}
                        stages_dates = list(stages_dic.keys())
                        stages_dates.sort()
                        # Если перед статусом Завершенная нет статуса Выполненная - добавляем
                        for k, stage_date in enumerate(stages_dates):
                            if k:
                                if stages_dic[stage_date]['status'] == 'Завершенная' \
                                                and stages_dic[stages_dates[k-1]]['status'] != 'Выполненная':
                                    stages_dic[stage_date - timedelta(microseconds=1)] = {
                                        'project': stages_dic[stage_date]['project'],
                                        'status': 'Выполненная',
                                        'user': stages_dic[stage_date]['user']}
                        stages = []
                        last_datetime = stages_dates[0]
                        stages_key = -1
                        for k, stage_date in enumerate(stages_dates):  # Еще раз фильтруем дубли и рассчитываем время
                            if k:
                                if stages_dic[stage_date]['status'] == stages_dic[stages_dates[k-1]]['status'] and \
                                        stages_dic[stage_date]['project'] == stages_dic[stages_dates[k-1]]['project'] \
                                        and stage_date >= stages_dates[k-1]:
                                    pass                    # Пропускаем если дублируются статус и проект, а время >=
                                else:
                                    stages_key += 1
                                    stages.append([stage_date, stages_dic[stage_date]['project'],
                                                   stages_dic[stage_date]['status'], stages_dic[stage_date]['user']])
                                    # Добавляем время длительности прошлого этапа, в зависимости от линии ТП
                                    if stages[stages_key - 1][1] == 'ТП 1 линия' and \
                                            stages[stages_key - 1][2] != 'ТехПоддержка Финброкер' and \
                                            stages[stages_key - 1][2] != 'Работа с Поставщиком':
                                        stages[stages_key - 1].append(self.workTime(last_datetime, stage_date,
                                                                      alldays=True))
                                    else:
                                        stages[stages_key - 1].append(self.workTime(last_datetime, stage_date,
                                                                      alldays=False))
                                    last_datetime = stage_date
                            else:
                                stages_key += 1             # Первую стадию берем без длительности прошлого этапа
                                stages.append([stage_date, stages_dic[stage_date]['project'],
                                               stages_dic[stage_date]['status'],
                                               stages_dic[stage_date]['user']])
                        stages[stages_key].append(timedelta(microseconds=0)) # Время длительности последнего этапа
                        for k, stage in enumerate(stages):
                            evolution += '➡' + stage[1] + '-' + stage[2] + '(' + \
                                             str(round(stage[4].total_seconds() / 60)) + ')'
                            for converse in CONVERSION:
                                evolution = evolution.replace(converse, CONVERSION[converse])
                        evolution = evolution.strip('➡')
                        last_project = ''
                        work_datetime = {'tp1NEW':'','tp1':'','itNEW':'','it':'','vndrNEW':'','vndr':'','fbNEW':'',
                                         'fb':'','rkoNEW':'','rko':'','done':'','completed':''}
                        work_operator = {'tp1':'','it':'','vndr':'','fb':'','rko':'','done':'','completed':''}
                        work_timedelta = {'tp1NEW':timedelta(milliseconds=1),'tp1':timedelta(milliseconds=1),
                                          'itNEW':timedelta(milliseconds=1), 'it':timedelta(milliseconds=1),
                                          'vndrNEW':timedelta(milliseconds=1), 'vndr':timedelta(milliseconds=1),
                                          'fbNEW':timedelta(milliseconds=1), 'fb':timedelta(milliseconds=1),
                                          'rkoNEW':timedelta(milliseconds=1), 'rko':timedelta(milliseconds=1),
                                          'done':timedelta(milliseconds=1)}
                        tasks[int(row[colnames.index('№ задачи')])] = stages
                        last_stage = ''
                        last_voronka = ''
                        last_datetime = datetime(1971,1,1,1,0,0)
                        timedelta_tp_expenses = timedelta(milliseconds=1)   # stage =
                        for k, stage in enumerate(stages):          # [ДатаВремя, Проект, Статус, Исполнитель, Период]
                            if stage[1] == 'ТП 1 линия' and stage[2] == 'Новая':
                                if not work_datetime['tp1NEW']:
                                    work_datetime['tp1NEW'] = stage[0]
                                work_timedelta['tp1NEW'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == 'ТП 1 линия' and stage[2] == 'В работе':
                                if not work_datetime['tp1']:
                                    work_datetime['tp1'] = stage[0]
                                    work_operator['tp1'] = stage[3]
                                work_timedelta['tp1'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == 'ТП 1 линия' and stage[2] == 'Работа с Поставщиком':
                                if not work_datetime['vndr']:
                                    work_datetime['vndr'] = stage[0]
                                    work_operator['vndr'] = stage[3]
                                work_timedelta['vndr'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == 'ТП Финброкер' and stage[2] == 'Новая':
                                if not work_datetime['fbNEW']:
                                    work_datetime['fbNEW'] = stage[0]
                                work_timedelta['fbNEW'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == 'ТП Финброкер' and stage[2] == 'В работе':
                                if not work_datetime['fb']:
                                    work_datetime['fb'] = stage[0]
                                    work_operator['fb'] = stage[3]
                                work_timedelta['fb'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == 'Заявки для РКО' and stage[2] == 'Новая':
                                if not work_datetime['rkoNEW']:
                                    work_datetime['rkoNEW'] = stage[0]
                                work_timedelta['rkoNEW'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == 'Заявки для РКО' and stage[2] == 'В работе':
                                if not work_datetime['rko']:
                                    work_datetime['rko'] = stage[0]
                                    work_operator['rko'] = stage[3]
                                work_timedelta['rko'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == '_Запросы на техподдержку' and stage[2] == 'Новая':
                                if not work_datetime['itNEW']:
                                    work_datetime['itNEW'] = stage[0]
                                work_timedelta['itNEW'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[1] == '_Запросы на техподдержку' and (
                                    stage[2] == 'В работе' or
                                    stage[2] == 'В очереди' or
                                    stage[2] == 'Дать ответ'):
                                if not work_datetime['it']:
                                    work_datetime['it'] = stage[0]
                                    work_operator['it'] = stage[3]
                                work_timedelta['it'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[2] == 'Выполненная':
                                if not work_datetime['done']:
                                    work_datetime['done'] = stage[0]
                                    work_operator['done'] = stage[3]
                                work_timedelta['done'] += stage[4]
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                            elif stage[2] == 'Завершенная':
                                work_datetime['completed'] = stage[0]
                                work_operator['completed'] = stage[3]
                            else:
                                last_stage = stage[1] + ' - ' + stage[2]
                                last_datetime = stage[0]
                        for delta in work_timedelta:
                            if delta != 'done':
                                timedelta_tp_expenses += work_timedelta[delta]
                        data_sozd_dec = stages[0][0].date().toordinal() - datetime(1900,1,1).toordinal() + 2
                        time_sozd = stages[0][0].time()
                        time_sozd_dec = int(stages[0][0].time().strftime('%H')) / 24 + \
                                        int(stages[0][0].time().strftime('%M')) / 1440 + \
                                        int(stages[0][0].time().strftime('%S')) / 86400
                        ws_rez.append([str(i), int(row[colnames.index('№ задачи')]),
                                       row[colnames.index('Название задачи')], row[colnames.index('Постановщик')],
                                       stages[0][1], stages[0][0], round(work_timedelta['tp1NEW'].total_seconds()/60),
                                       round(work_timedelta['tp1'].total_seconds()/60), work_operator['tp1'],
                                       round(work_timedelta['itNEW'].total_seconds()/60),
                                       round(work_timedelta['it'].total_seconds()/60), work_operator['it'],
                                       round(work_timedelta['vndr'].total_seconds()/60), work_operator['vndr'],
                                       round(work_timedelta['fbNEW'].total_seconds()/60),
                                       round(work_timedelta['fb'].total_seconds()/60), work_operator['fb'],
                                       round(work_timedelta['rkoNEW'].total_seconds()/60),
                                       round(work_timedelta['rko'].total_seconds()/60), work_operator['rko'],
                                       work_datetime['done'],
                                       round(timedelta_tp_expenses.total_seconds()/60), work_operator['done'],
                                       work_datetime['completed'], last_stage, evolution,
                                       row[colnames.index('Внутренний продукт')],
                                       row[colnames.index('Канал продаж')],
                                       row[colnames.index('Источник добавления')],
                                       row[colnames.index('Источники проблем')],
                                       row[colnames.index('Краткое описание проблемы')], data_sozd_dec, time_sozd_dec])
            self.progressBar['value'] = ws.max_row
            self.update()
            wb_rez.save(os.path.join(os.path.dirname(self.xlsx),
                                     '.'.join(os.path.basename(self.xlsx).replace(' ', '_').split('.')[:-1])) +
                        '_обр.xlsx')
            return


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    root = Tk()
    ex = Example(root)
    #root.geometry("490x127+100+100")
    root.mainloop()
