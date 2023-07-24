# Определение региона и оператора по телефону

import warnings
import openpyxl
import os
import sys
from datetime import datetime, timedelta, time, date
from string import digits
import csv

from tkinter import Tk, Frame, Button, Menu, filedialog, BooleanVar, Checkbutton, Radiobutton, ttk
from tkinter import messagebox as mbox
from lib import l, format_phone

MAX_BLANK_LINES = 5

class Example(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.parent = parent
        self.initUI()

    def initUI(self):
        self.parent.title("Определение региона и оператора по телефону v 0.4 от 20.10.2021")
        self.pack()
        # Главное меню .
        menubar = Menu(self.parent, font=('Calibri 10'))
        self.parent.config(menu=menubar)
        # Раздел меню, без открепления
        self.fileMenu = Menu(menubar, tearoff=0, font=('Calibri 10'))
        # Пункты раздела меню
        self.fileMenu.add_command(label="Открыть список телефонов (Excel)", command=self.onOpen)
        self.fileMenu.add_separator()
        self.fileMenu.add_command(label="Выход", command=self.onExit)
        # Присоединяем раздел каскадом
        menubar.add_cascade(label="Файл", menu=self.fileMenu)

        btnOpen = Button(self, text="Открыть список телефонов (Excel)", command=self.onOpen, font=('Calibri 10'))
        btnOpen.grid(padx=5, pady=5, row=0, column=0)

        self.progressBar = ttk.Progressbar(self, orient="horizontal", mode="determinate", maximum=100, value=0,
                                           length=200)
        self.progressBar.grid(padx=5, pady=5, row=1, column=0)#, columnspan=2)

    def onExit(self):
        #mbox.showerror('Ошибка в файле' +  self.xlsx, 'Нет столбца с... ')
        sys.exit()

    def onOpen(self):
        self.progressBar['value'] = 0
        self.update()
        # Сортируем так, чтобы самые свежие по дате создания были вначале
        files = sorted(filter(os.path.isfile, os.listdir('.')), key=os.path.getmtime, reverse=True)
        csv_file = ''
        for file in files:
            if file.endswith('.csv'):           # Выбираем первый попавшийся .csv
                csv_file = file
                break
        operators = {}
        regions = {}
        deltas = {}
        if csv_file:
            with open(csv_file, encoding='utf-8') as r_file:
                file_reader = csv.DictReader(r_file, delimiter=";")
                last_phone = 78999999999
                for row in file_reader:
                    start_phone = 70000000000 + int(row['АВС/ DEF']) * 10000000 + int(row['От'])
                    end_phone = 70000000000 + int(row['АВС/ DEF']) * 10000000 + int(row['До'])
                    if last_phone + 1 < start_phone:
                        operators[last_phone + 1] = '---не определено в csv файле'
                        regions[last_phone + 1] = '---не определено в csv файле'
                        deltas[last_phone + 1] = '---не определено в csv файле'
                        operators[start_phone] = row['Оператор']
                        regions[start_phone] = row['Регион']
                        deltas[start_phone] = row['Разница во времени с Москвой']
                    elif last_phone + 1 == start_phone:
                        operators[start_phone] = row['Оператор']
                        regions[start_phone] = row['Регион']
                        deltas[start_phone] = row['Разница во времени с Москвой']
                    else:
                        mbox.showerror('Ошибка в файле' +  csv_file, 'Нарушен порядок следования номеров')
                        return
                    last_phone = end_phone
        operators[80000000000] = '---не определено в csv файле'
        regions[80000000000] = '---не определено в csv файле'
        deltas[80000000000] = '---не определено в csv файле'
        phones_keys = list(operators.keys())
        phones_keys.sort()

        dlg = filedialog.Open(self, filetypes=[('Файлы Excel 2007-2020', '*.xlsx')])
        xlsx = dlg.show()
        if str(xlsx).find('.xlsx') > -1:   # Проверяем то, что и так понятно
            wb = openpyxl.load_workbook(filename=xlsx, read_only=True)
            ws = wb[wb.sheetnames[0]]
            wb_rez = openpyxl.Workbook(write_only=True)
            blank_lines = 0
            self.progressBar['maximum'] = ws.max_row
            self.progressBar['value'] = 0
            self.update()
            ws_rez = wb_rez.create_sheet('Отчет работы всех линий ТП')
            ws_rez.append(['№ п/п', 'Мобильный телефон', 'Оператор', 'Регион', 'Разница во времени с Москвой'])
            j = 0
            for i, row in enumerate(ws.values):                 # Идем по строчкам отчета
                self.progressBar['value'] = i
                self.update()
                if not format_phone(row[0]):   # Если в столбце 0 не телефон
                    blank_lines += 1
                    if blank_lines > MAX_BLANK_LINES:
                        break
                elif l(format_phone(row[0])) < 79000000000:
                    blank_lines = 0
                    j += 1
                    ws_rez.append([j, format_phone(row[0]), '---телефон не 79хх', '---телефон не 79хх',
                                   '---телефон не 79хх'])
                else:
                    blank_lines = 0
                    last_phones_key = 79000000000
                    for phones_key in phones_keys:
                        if phones_key > format_phone(row[0]):
                            j += 1
                            ws_rez.append([j, format_phone(row[0]), operators[last_phones_key],
                                           regions[last_phones_key], deltas[last_phones_key]])
                            break
                        last_phones_key = phones_key
            wb_rez.save(os.path.splitext(xlsx)[0] + '_обр.xlsx')
            self.progressBar['value'] = 0
            self.update()


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    root = Tk()
    ex = Example(root)
    #root.geometry("490x127+100+100")
    root.mainloop()
