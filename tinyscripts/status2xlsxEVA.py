# Обновление статусов прозвона Рокет в xlsx из mongo. Нужен файл с настройками
# подключения anketa.ini. Закинуть экселевский документ в папку со скриптом.
# Скрипт проверяет 3й столбец таблицы. Из 3ей колонки берет remote_id, в 4ю ложит результат.
# Выполнить скрипт, результат запишется в ту же таблицу.

import sys, argparse
from _datetime import datetime, timedelta, date
import time
import os
from mysql.connector import MySQLConnection, Error
from collections import OrderedDict
import openpyxl
from pymongo import MongoClient
import psycopg2

from lib import read_config, s, l

cfg = read_config(filename='anketa.ini', section='Mongo')
# подключ к монге
conn = MongoClient('mongodb://' + cfg['user'] + ':' + cfg['password'] + '@' + cfg['ip'] + ':' + cfg['port'] + '/'
                    + cfg['db'])
# выбираем базу данных
db = conn.saturn_v
# выбираем коллекцию документов
collection = db.Products
path = "./"
file_list = os.listdir(path)
full_list = [os.path.join(path, i) for i in file_list if i.endswith('.xlsx')]
xlsxs = sorted(full_list, key = os.path.getmtime)
f=open("guru99.txt", "a+")


# for each xls document in a list
for xlsx in xlsxs:
    wb = openpyxl.load_workbook(filename=xlsx)
    sheet = wb[wb.sheetnames[2]]
    # столбец с айди
    column_id = 2
    # счетчик, в то же время номер строки
    kk = 1
    # сборище месаджей в будущем
    pack = []
    for i, row in enumerate(sheet.values):        
        msg = ""
        # забирает айдишник из строки
        str_excel_id = str(row[column_id])
        print("_______")
        print("excel id " + str_excel_id)
        #print("number of excel string " + str(i))
        for doc in collection.find({ "remote_id": str_excel_id }, 
                                            { 'callcenter_meta.message': 1 }):           
            pack.append(doc)
            #print("message from mongo " + str(doc))
            print(kk)
            if 'callcenter_meta' in doc and 'message' in doc["callcenter_meta"]:
                print(doc["callcenter_meta"]["message"])
                msg = doc["callcenter_meta"]["message"]
            else:
                print("no message")
            print("\n")
        # пропустить первую строку, там заголовок
        if kk == 1:
            pass
        else:
            _ = sheet.cell(column=4, row=kk, value=msg)
        print(msg)
        kk += 1
    wb.save(filename = xlsx)
    print("хатова, праверяй!")
