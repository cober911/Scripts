# Свод РКО по Ф67 (из многих файлов xlsx в один большой сводный)

import warnings
import openpyxl
import os
import os.path
import sys
from datetime import datetime, timedelta, time, date
import xlrd

from tkinter import Tk, Frame, Button, Menu, filedialog, BooleanVar, Checkbutton, Radiobutton, ttk
from lib import l

MAX_BLANK_LINES = 5
COLNAMES = ['ИНН организации/ИП', 'Название организации/ИП', 'Дата формирования', 'Полное название банка', 'БИК банка',
            'Краткое название банка' , 'Дата открытия/дата предоставления права',
            'Дата закрытия/дата прекращения права', 'Состояние', 'Вид счета']

class Example(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.parent = parent
        self.initUI()

    def initUI(self):
        self.parent.title("Свод РКО по Ф67 v 0.4 от 09.11.2021")
        self.pack()
        # Главное меню .
        menubar = Menu(self.parent, font=('Calibri 10'))
        self.parent.config(menu=menubar)
        # Раздел меню, без открепления
        self.fileMenu = Menu(menubar, tearoff=0, font=('Calibri 10'))
        # Пункты раздела меню
        self.fileMenu.add_command(label="Открыть директорию с Ф67", command=self.onOpen)
        self.fileMenu.add_separator()
        self.fileMenu.add_command(label="Выход", command=self.onExit)
        # Присоединяем раздел каскадом
        menubar.add_cascade(label="Файл", menu=self.fileMenu)

        btnOpen = Button(self, text="Открыть директорию с Ф67", command=self.onOpen, font=('Calibri 10'))
        btnOpen.grid(padx=5, pady=5, row=0, column=0)

        self.progressBar = ttk.Progressbar(self, orient="horizontal", mode="determinate", maximum=100, value=0,
                                           length=200)
        self.progressBar.grid(padx=5, pady=5, row=1, column=0)#, columnspan=2)

    def onExit(self):
        #mbox.showerror('Ошибка в файле' +  self.xlsx, 'Нет столбца с... ')
        sys.exit()

    def onOpen(self):
        self.progressBar['value'] = 0
        self.update()
        dir4load = filedialog.askdirectory(initialdir='.')
        # Фильтруем файлы .xls и сортируем так, чтобы самые свежие по дате создания были вначале
        # xlss = sorted(filter(lambda x: x.endswith('.xls'), filter(os.path.isfile, os.listdir(dir4load))), key=os.path.getmtime, reverse=True)
        xlss = list(filter(lambda x: x.endswith('.xls'), os.listdir(dir4load)))
        wb_rez = openpyxl.Workbook(write_only=True)
        ws_rez = wb_rez.create_sheet('Свод')
        ws_rez.append(COLNAMES)
        wb_short_names = openpyxl.load_workbook(filename='сокращенное название банков.xlsx', read_only=True)
        ws_short_names = wb_short_names[wb_short_names.sheetnames[0]]
        blank_lines = 0
        full_names = {}
        short_names = {}
        for i, row in enumerate(ws_short_names.values):
            if not l(row[0]):
                blank_lines += 1
                if blank_lines > 10:
                    break
            else:
                blank_lines = 0
                if row[1]:
                    full_names[int(row[0])] = row[1]
                if row[2]:
                    short_names[int(row[0])] = row[2]
        wb_short_names.close()
        self.progressBar['maximum'] = len(xlss)
        self.progressBar['value'] = 0
        self.update()
        for i, xls in enumerate(xlss):
            wb = xlrd.open_workbook(os.path.join(dir4load, xls), on_demand=True)
            ws = wb.sheet_by_index(0)
            blank_lines = 0
            self.progressBar['value'] = i
            self.update()
            last_cell_0 = ''
            title_loading = True
            rez_dic = {}
            for j in range(ws.nrows):                 # Идем по строчкам Ф67
                row = ws.row_values(j, 0, ws.ncols)
                if not row[0] and not row[1]:   # Если столбцы 0 и 1 пустые > MAX_BLANK_LINES раз, то выходим
                    blank_lines += 1
                    if blank_lines > MAX_BLANK_LINES:
                        break
                else:
                    blank_lines = 0
                    if title_loading:
                        if str(row[0]).find('ИНН:') > -1 and str(row[0]).find('ОГРН') > -1:
                            rez_dic['ИНН организации/ИП'] = int(str(row[0]).split(': ')[1].split(',')[0])
                            rez_dic['Название организации/ИП'] = last_cell_0
                        elif str(row[0]).find('Дата формирования') > -1:
                            for data in str(row[0]).split(' '):
                                if data.find('.') > -1:
                                    if data.find('.', data.find('.') + 1) > -1:
                                        rez_dic['Дата формирования'] = datetime.strptime(data.strip().strip(),
                                                                                         '%d.%m.%Y')
                        elif str(row[0]) == 'КПП (организации)':
                            title_loading = False
                    else:
                        if str(row[0]).find('РегНом/НомФ: ') > -1:
                            # очищаем параметры предыдущего банка
                            rez_dic = {colname: rez_dic[colname] for k, colname in enumerate(COLNAMES) if k < 3}
                            rez_dic['Полное название банка'] = last_cell_0
                        elif str(row[0]).find('БИК : ') > -1:
                            rez_dic['БИК банка'] = int(str(row[0]).split(': ')[1])
                            if full_names.get(rez_dic['БИК банка']):
                                rez_dic['Полное название банка'] = full_names[rez_dic['БИК банка']]
                            if short_names.get(rez_dic['БИК банка']):
                                rez_dic['Краткое название банка'] = short_names[rez_dic['БИК банка']]
                        elif row[1]:
                            # очищаем параметры предыдущего счета
                            rez_dic = {colname: rez_dic[colname] for k, colname in enumerate(COLNAMES) if k < 6}
                            if str(row[2]).find('.') > -1:
                                if str(row[2]).find('.', str(row[2]).find('.') + 1) > -1:
                                    rez_dic['Дата открытия/дата предоставления права'] = datetime.strptime(
                                                                                str(row[2]).strip().strip(), '%d.%m.%Y')
                            if str(row[3]).find('.') > -1:
                                if str(row[3]).find('.', str(row[3]).find('.') + 1) > -1:
                                    rez_dic['Дата закрытия/дата прекращения права'] = datetime.strptime(
                                                                                str(row[3]).strip().strip(), '%d.%m.%Y')
                            rez_dic['Состояние'] = row[4]
                            rez_dic['Вид счета'] = row[5]
                            ws_rez.append(list([rez_dic.get(colname, '') for colname in COLNAMES]))
                    last_cell_0 = row[0]
        wb_rez.save(os.path.join(dir4load, datetime.now().strftime("%d-%m-%Y_%H-%M") + '_свод.xlsx'))
        self.progressBar['value'] = 0
        self.update()


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    root = Tk()
    ex = Example(root)
    #root.geometry("490x127+100+100")
    root.mainloop()
